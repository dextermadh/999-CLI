import os
from pathlib import Path
from typing import List, Union
from tools.vision_engine import VisionEngine

class LocalFileManager:
    def __init__(self, project_root: str):
        # Resolve to absolute path to prevent ".." directory traversal attacks
        self.root = Path(project_root).resolve()
        
        # Initialize Vision Engine with common Tesseract paths
        tess_path = None
        possible_paths = [
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Users\{}\AppData\Local\Tesseract-OCR\tesseract.exe".format(os.getlogin())
        ]
        for p in possible_paths:
            if os.path.exists(p):
                tess_path = p
                break
        self.vision = VisionEngine(tesseract_path=tess_path)

    def _is_safe_path(self, path: Union[str, Path]) -> bool:
        """Ensures the agent is not trying to access files outside the project root."""
        target_path = Path(path).resolve()
        return self.root in target_path.parents or target_path == self.root

    def generate_diff(self, file_path: str, old_content: str, new_content: str) -> str:
        import difflib
        old_lines = old_content.splitlines(keepends=True)
        new_lines = new_content.splitlines(keepends=True)
        diff = difflib.unified_diff(
            old_lines, new_lines,
            fromfile=f"a/{file_path}",
            tofile=f"b/{file_path}",
            n=3
        )
        return "".join(diff)

    def preview_write_file(self, file_path: str, content: str) -> str:
        target_file = (self.root / file_path).resolve()
        if not self._is_safe_path(target_file):
            return "Error: Access Denied."
        old_content = target_file.read_text(encoding="utf-8") if target_file.exists() else ""
        return self.generate_diff(file_path, old_content, content)

    def preview_patch_file(self, file_path: str, search_string: str, replace_string: str) -> str:
        target_file = (self.root / file_path).resolve()
        if not self._is_safe_path(target_file):
            return "Error: Access Denied."
        if not target_file.exists():
            return f"Error: File {file_path} does not exist."
        content = target_file.read_text(encoding="utf-8")
        if search_string not in content:
            return f"Error: Search string not found in {file_path}."
        new_content = content.replace(search_string, replace_string, 1)
        return self.generate_diff(file_path, content, new_content)

    def list_files(self, subdir: str = ".") -> List[str]:
        """Lists files in a directory, respecting the sandbox."""
        target_dir = (self.root / subdir).resolve()
        
        if not self._is_safe_path(target_dir):
            return ["Error: Access Denied. Path is outside sandbox."]
        
        try:
            return [str(p.relative_to(self.root)) for p in target_dir.iterdir() if p.is_file()]
        except Exception as e:
            return [f"Error: {str(e)}"]

    def read_file(self, file_path: str, max_lines: int = 300) -> str:
        """Reads content of a local file, supporting text, PDF, Word, and Excel.
        For code/text files, intelligently truncates to max_lines to prevent
        context window overflow. Use view_file_lines for specific ranges.
        """
        target_file = (self.root / file_path).resolve()
        
        if not self._is_safe_path(target_file):
            return "Error: Access Denied."
            
        if not target_file.exists():
            return f"Error: File {file_path} not found."

        ext = target_file.suffix.lower()
        
        try:
            if ext in ['.txt', '.py', '.js', '.ts', '.tsx', '.jsx', '.html', '.css', '.md', '.json', '.yaml', '.yml', '.sql', '.sh', '.go', '.rs', '.java', '.c', '.cpp', '.h', '.toml', '.cfg', '.ini', '.bat', '.ps1']:
                content = target_file.read_text(encoding="utf-8")
                lines = content.split('\n')
                total_lines = len(lines)
                
                # Intelligent truncation for large files
                if total_lines > max_lines:
                    head_count = max_lines - 50  # e.g., 250 lines from top
                    tail_count = 50
                    omitted = total_lines - head_count - tail_count
                    head = '\n'.join(lines[:head_count])
                    tail = '\n'.join(lines[-tail_count:])
                    return (
                        f"{head}\n\n"
                        f"... [{omitted} lines omitted — file has {total_lines} total lines. "
                        f"Use view_file_lines for specific ranges.] ...\n\n"
                        f"{tail}"
                    )
                return content
            
            elif ext == '.pdf':
                from pypdf import PdfReader
                reader = PdfReader(target_file)
                text = []
                for page in reader.pages:
                    text.append(page.extract_text())
                
                # Perform OCR on embedded images
                ocr_text = self.vision.extract_and_ocr_pdf(target_file)
                if ocr_text:
                    text.append("\n[DEEP SCAN: OCR TEXT FROM IMAGES]\n" + ocr_text)
                
                full_text = "\n".join(text)
                # Truncate very large documents
                if len(full_text) > 15000:
                    return full_text[:15000] + f"\n\n... [Document truncated at 15000 chars. Total: {len(full_text)} chars.]"  
                return full_text
                
            elif ext == '.docx':
                import docx
                doc = docx.Document(target_file)
                text = [para.text for para in doc.paragraphs]
                
                # Perform OCR on embedded images
                ocr_text = self.vision.extract_and_ocr_docx(target_file)
                if ocr_text:
                    text.append("\n[DEEP SCAN: OCR TEXT FROM IMAGES]\n" + ocr_text)
                    
                return "\n".join(text)
                
            elif ext in ['.xlsx', '.xls']:
                import openpyxl
                if ext == '.xlsx':
                    wb = openpyxl.load_workbook(target_file, data_only=True)
                    output = []
                    for sheet in wb.sheetnames:
                        output.append(f"Sheet: {sheet}")
                        ws = wb[sheet]
                        for row in ws.iter_rows(values_only=True):
                            output.append("\t".join([str(c) if c is not None else "" for c in row]))
                    return "\n".join(output)
                else:
                    return "Error: .xls (Legacy Excel) requires xlrd. Converting to .xlsx is recommended."

            elif ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp']:
                from PIL import Image
                with Image.open(target_file) as img:
                    metadata = f"Image Metadata:\nFormat: {img.format}\nSize: {img.size}\nMode: {img.mode}"
                    ocr_text = self.vision.ocr_image(img)
                    if ocr_text:
                        metadata += f"\n\nOCR TEXT CONTENT:\n{ocr_text}"
                    return metadata

            elif ext == '.csv':
                import csv
                output = []
                with open(target_file, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        output.append(",".join(row))
                return "\n".join(output)

            else:
                # Default to text if unknown
                return target_file.read_text(encoding="utf-8")
                
        except Exception as e:
            return f"Error reading {ext} file: {str(e)}"

    def write_file(self, file_path: str, content: str) -> str:
        """Writes content to a file. Useful for the agent creating code."""
        target_file = (self.root / file_path).resolve()
        
        if not self._is_safe_path(target_file):
            return "Error: Access Denied."
            
        try:
            # Ensure parent directories exist
            target_file.parent.mkdir(parents=True, exist_ok=True)
            target_file.write_text(content, encoding="utf-8")
            return f"Successfully wrote to {file_path}"
        except Exception as e:
            return f"Error writing file: {str(e)}"

    def patch_file(self, file_path: str, search_string: str, replace_string: str) -> str:
        """Replaces a specific block of text in a file with new text."""
        target_file = (self.root / file_path).resolve()
        
        if not self._is_safe_path(target_file):
            return "Error: Access Denied."
            
        if not target_file.exists():
            return f"Error: File {file_path} does not exist."
            
        try:
            content = target_file.read_text(encoding="utf-8")
            if search_string not in content:
                return f"Error: Search string not found in {file_path}. Ensure exact match including whitespace."
                
            new_content = content.replace(search_string, replace_string, 1) # Replace first occurrence
            target_file.write_text(new_content, encoding="utf-8")
            return f"Successfully patched {file_path}"
        except Exception as e:
            return f"Error patching file: {str(e)}"

    def delete_file(self, file_path: str) -> str:
        """Safely deletes a file within the workspace."""
        target_file = (self.root / file_path).resolve()
        if not self._is_safe_path(target_file):
            return "Error: Access Denied."
        try:
            if target_file.is_file():
                target_file.unlink()
                return f"Successfully deleted file {file_path}"
            elif target_file.is_dir():
                import shutil
                shutil.rmtree(target_file)
                return f"Successfully deleted directory {file_path}"
            return f"Error: {file_path} not found"
        except Exception as e:
            return f"Error deleting: {str(e)}"

    def move_file(self, source_path: str, dest_path: str) -> str:
        """Moves or renames a file/directory."""
        src = (self.root / source_path).resolve()
        dst = (self.root / dest_path).resolve()
        if not self._is_safe_path(src) or not self._is_safe_path(dst):
            return "Error: Access Denied."
        try:
            if not src.exists():
                return f"Error: Source {source_path} does not exist."
            dst.parent.mkdir(parents=True, exist_ok=True)
            import shutil
            shutil.move(str(src), str(dst))
            return f"Successfully moved {source_path} to {dest_path}"
        except Exception as e:
            return f"Error moving: {str(e)}"

    def create_dir(self, dir_path: str) -> str:
        """Creates a directory."""
        target_dir = (self.root / dir_path).resolve()
        if not self._is_safe_path(target_dir):
            return "Error: Access Denied."
        try:
            target_dir.mkdir(parents=True, exist_ok=True)
            return f"Successfully created directory {dir_path}"
        except Exception as e:
            return f"Error creating directory: {str(e)}"

    def get_file_info(self, file_path: str) -> str:
        """Returns metadata about a file."""
        target_file = (self.root / file_path).resolve()
        if not self._is_safe_path(target_file):
            return "Error: Access Denied."
        try:
            if not target_file.exists():
                return f"Error: {file_path} does not exist."
            stats = target_file.stat()
            size = stats.st_size
            import datetime
            modified = datetime.datetime.fromtimestamp(stats.st_mtime).isoformat()
            info = f"Size: {size} bytes\nLast Modified: {modified}"
            if target_file.is_file():
                try:
                    lines = sum(1 for _ in open(target_file, 'r', encoding='utf-8'))
                    info += f"\nLines: {lines}"
                except:
                    info += "\nLines: Binary or unreadable"
            return info
        except Exception as e:
            return f"Error getting info: {str(e)}"